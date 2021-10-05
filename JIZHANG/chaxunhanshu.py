import tkinter as tk
from tkinter import ttk
import datetime
import datepy
import openpyxl

#!记账页点查询按钮转入界面


class chaxunhanshu(tk.Tk):

    # ! 初始化
    def __init__(self):
        super().__init__()

        self.title("个人财务系统查询")
        self.geometry("+800+250")
        self.resizable(0, 0)
        self.jiemian()

    #!明细下拉列表值
    def minxihan(self):
        x = []
        wb = openpyxl.load_workbook('Z.xlsx')
        ws = wb['收入明细']
        for i in range(2, ws.max_row):
            list1 = ws.cell(i, 3).value
            x.append(list1)
        x = list(set(x))
        return x

    # ! 建下拉列表
    def lianyueru(self, value, x, y, z):
        self.lab_data1 = ttk.Combobox(self.frame2, textvariable=value, width=x)
        self.lab_data1.grid(row=0, column=y)
        self.lab_data1['values'] = z
        return self.lab_data1

    # ! 建显示标签
    def Lab(self, f, a, b, c, l, x, y):
        self.lab = tk.Label(f, text=a, bg=None,
                            relief=None, font=(b, c), width=l)
        self.lab.grid(row=x, column=y, padx=2, pady=2)
        return self.lab

    #! 建按钮
    def bnt(self, f, a, b, c, s, w, h, d, x, y, z, t):
        but = tk.Button(f, text=a, font=(b, c), bg=s,
                        width=w, height=h, command=d)
        but.grid(row=x, column=y, padx=z, pady=t)
        return but

    #!获取下拉列表中的值
    def queren(self):
        print(self.lab_data1['values'][self.lab_data1.current()])

    #!向前按钮
    def befor1(self):
        time1 = datetime.datetime.strptime(
            self.yeras.get(), '%Y-%m-%d').date()  # !把获取到的COMBOBOX中的初始值转化为DATE表格
        self.yeras['values'] = datepy.datelist(
            (time1-datetime.timedelta(9)), self.yeras.get())  # !调用datepy模块，生成新的COMBOBOX中的值（前10天）
        self.yeras.current(0)  # !重新赋初始值给COMBOBOX

    def befor2(self):
        time1 = datetime.datetime.strptime(
            self.yeras1.get(), '%Y-%m-%d').date()  # !把获取到的COMBOBOX中的初始值转化为DATE表格
        self.yeras1['values'] = datepy.datelist(
            (time1-datetime.timedelta(9)), self.yeras1.get())  # !调用datepy模块，生成新的COMBOBOX中的值（前10天）
        self.yeras1.current(0)

    #! 向后按钮
    def after1(self):
        time1 = datetime.datetime.strptime(
            self.yeras.get(), '%Y-%m-%d').date()  # !把获取到的COMBOBOX中的初始值转化为DATE表格
        self.yeras['values'] = datepy.datelist(self.yeras.get(
        ), (time1+datetime.timedelta(9)))  # !调用datepy模块，生成新的COMBOBOX中的值（前10天）
        self.yeras.current(9)  # !重新赋初始值给COMBOBOX

    def after2(self):
        time1 = datetime.datetime.strptime(
            self.yeras1.get(), '%Y-%m-%d').date()  # !把获取到的COMBOBOX中的初始值转化为DATE表格
        self.yeras1['values'] = datepy.datelist(self.yeras1.get(
        ), (time1+datetime.timedelta(9)))  # !调用datepy模块，生成新的COMBOBOX中的值（前10天）
        self.yeras1.current(9)

    #!确认按钮函数
    def ok(self):
        n=0
        print(self.yeras.get())
        print(self.yeras1.get())
        date = datetime.datetime.strptime(str(self.yeras.get()), '%Y-%m-%d')
        date1 = datetime.datetime.strptime(str(self.yeras1.get()), '%Y-%m-%d')
        print(type(date))
        # print(type(date1))
        # print(type(self.yeras.get()))
        print(date,date1)
        wb = openpyxl.load_workbook('Z.xlsx')
        ws = wb['收入明细']
        for i in range(2, ws.max_row):
            list1 = ws.cell(i, 1).value
            list1=str(list1)
            print(type(list1))
            print(list1)
            qq= list1.split("\00")[0]
            list_1=datetime.datetime.strptime(qq, '%Y-%m-%d')
            # print(type(list_1))

            # if list_1>date and list_1<date1:
            #     a=ws.cell(i, 1).value
            #     b=ws.cell(i, 2).value
            #     c=ws.cell(i, 3).value
            #     d=ws.cell(i, 4).value
            #     self.Lab(self.frame3, a, '楷体', 12, 14, n, 0)
            #     self.Lab(self.frame3, b, '楷体', 12, 8, n, 1)
            #     self.Lab(self.frame3, c, '楷体', 12, 10, n, 2)
            #     self.Lab(self.frame3, d, '楷体', 12, 36, n, 3)
            #     n+=1

    # ! 查询界面
    def jiemian(self):

        minxihan = self.minxihan()
        self.frame1 = tk.Frame(bg="#e1f5c4", relief="sunken", bd=1)
        self.frame1.pack(fill=tk.X)
        self.frame2 = tk.Frame(bg="#e1f5c4", relief="sunken", bd=1)
        self.frame2.pack(fill=tk.X)
        self.frame3 = tk.Frame(bg="#e1f5c4", relief="sunken", bd=1)
        self.frame3.pack(fill=tk.X,)

        self.S1 = self.Lab(self.frame1, '总计', '楷体', 24, None, 0, 0)
        self.S2 = self.Lab(self.frame1, '400', '楷体', 24, None, 0, 1)

        self.S3 = self.Lab(self.frame2, '日期', '楷体', 24, None, 0, 0)
        self.but1 = self.bnt(self.frame2, '<', '楷体', '10',
                             'green', 1, 0, self.befor1, 0, 1, 0, 1)
        self.yeras = self.lianyueru(tk.StringVar(), 10, 2, datepy.datelist(
            (datetime.date.today()-datetime.timedelta(9)), datetime.date.today()))
        self.yeras.current(0)
        self.but2 = self.bnt(self.frame2, '>', '楷体', '10',
                             'green', 1, 1, self.after1, 0, 3, 1, 1)
        self.S4 = self.Lab(self.frame2, '--', '楷体', 24, None, 0, 4)
        self.but3 = self.bnt(self.frame2, '<', '楷体', '10',
                             'green', 1, 1, self.befor2, 0, 5, 1, 1)
        self.yeras1 = self.lianyueru(tk.StringVar(), 10, 6, datepy.datelist(
            (datetime.date.today()-datetime.timedelta(9)), datetime.date.today()))
        self.yeras1.current(9)
        self.but4 = self.bnt(self.frame2, '>', '楷体', '10',
                             'green', 1, 1, self.after2, 0, 7, 1, 1)
        self.S5 = self.Lab(self.frame2, '明细', '楷体', 24, None, 0, 8)
        self.minxi = self.lianyueru(tk.StringVar(), 8, 9, minxihan)
        self.minxi.current(0)
        self.but5 = self.bnt(self.frame2, '确认', '楷体', '14',
                             'green', 4, 1, self.ok, 0, 10, 1, 1)
        # print(len(minxihan))
        self.Lab(self.frame3, ' 日期', '楷体', 12, 14, 0, 0)
        self.Lab(self.frame3, '金额', '楷体', 12, 8, 0, 1)
        self.Lab(self.frame3, '明细', '楷体', 12, 10, 0, 2)
        self.Lab(self.frame3, '备注', '楷体', 12, 36, 0, 3)


#! 主程序
if __name__ == "__main__":

    bpp = chaxunhanshu()
    bpp.mainloop()
