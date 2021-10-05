import tkinter as tk
from PIL import Image, ImageTk
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import chaxunhanshu
import datetime


# 注册界面
class zhuce(tk.Tk):

    # 初始化
    def __init__(self):
        super().__init__()
        self.title("个人财务系统注册")
        self.geometry("+800+250")
        self.resizable(0, 0)
        self.zhuce1()

    # 注册界面中点击注册函数

    def zhuce1_jilu(self):
        # self.frame1_zhuce.pack(padx=20, pady=5, fill=tk.X)
        zhuce1_jilu_name = self.txt_username_zhuce.get()
        zhuce1_jilu_password = self.txt_password_zhuce.get()
        zhuce1_jilu_password1 = self.txt_password_zhuce1.get()
        if zhuce1_jilu_name != '' and zhuce1_jilu_password != '' and zhuce1_jilu_password1 != '':
            if zhuce1_jilu_password == zhuce1_jilu_password1:
                wb = Workbook()
                wb1 = wb.create_sheet("C", 0)
                wb1['A1'] = zhuce1_jilu_name
                wb1['B1'] = zhuce1_jilu_password
                wb1.sheet_properties.tabColor = "1072BA"
                readname = zhuce1_jilu_name
                wb2 = wb.create_sheet('收入明细', 1)
                wb2['A1'] = '日期'
                wb2['B1'] = '金额'
                wb2['C1'] = '明细'
                wb2['D1'] = '备注'
                # wb2.sheet_properties.tabColor='green'
                wb3 = wb.create_sheet('支出明细', 2)
                wb3['A1'] = '日期'
                wb3['B1'] = '金额'
                wb3['C1'] = '明细'
                wb3['D1'] = '备注'
                # wb3.sheet_properties.tabColor='red'
                wb.save('D:\python\PYTHON\JIZHANG\Z.xlsx')
                self.destroy()
            else:
                self.frame1_zhuce.pack()
                self.lbl_baocuo1.config(
                    text='您输入的两次密码不相同', font=('微软雅黑', '18'), fg='red')
                self.txt_username_zhuce.delete(0, tk.END)
                self.txt_password_zhuce.delete(0, tk.END)
                self.txt_password_zhuce1.delete(0, tk.END)
                self.frame1_zhuce.after(2000, self.frame1_zhuce.pack_forget)
        else:
            self.frame1_zhuce.pack()
            self.lbl_baocuo1.config(
                text='账户或密码不能为空', font=('微软雅黑', '18'), fg='red')
            self.frame1_zhuce.after(2000, self.frame1_zhuce.pack_forget)

    # 生成注册界面

    def zhuce1(self):
        self.frame_zhuce = tk.Frame(self)

        self.lbl_username_zhuce = tk.Label(
            self.frame_zhuce, text="用户名", anchor=tk.W, width=40)
        self.lbl_username_zhuce.pack(fill=tk.X)

        self.txt_username_zhuce = tk.Entry(self.frame_zhuce)
        self.txt_username_zhuce.pack(fill=tk.X)

        self.lbl_password_zhuce = tk.Label(
            self.frame_zhuce, text="密码", anchor=tk.W)
        self.lbl_password_zhuce.pack(fill=tk.X)

        self.txt_password_zhuce = tk.Entry(self.frame_zhuce, show="*")
        self.txt_password_zhuce.pack(fill=tk.X)

        self.lbl_password_zhuce1 = tk.Label(
            self.frame_zhuce, text="确认密码", anchor=tk.W)
        self.lbl_password_zhuce1.pack(fill=tk.X)

        self.txt_password_zhuce1 = tk.Entry(self.frame_zhuce, show="*")
        self.txt_password_zhuce1.pack(fill=tk.X)

        self.btn_zhuce = tk.Button(
            self.frame_zhuce, text="注册", width="16", command=self.zhuce1_jilu)
        self.btn_zhuce.pack(pady=10)

        self.frame_zhuce.pack(padx=20, pady=5, fill=tk.X)

        self.frame1_zhuce = tk.Frame(self)

        self.lbl_baocuo1 = tk.Label(
            self.frame1_zhuce, text="", anchor=tk.CENTER)
        self.lbl_baocuo1.pack()

        self.frame1_zhuce.pack_forget()


# 录入界面
class Financial_system(tk.Tk):

    # 初始化
    def __init__(self):
        super().__init__()
        self.title("财务系统")
        self.geometry("+800+250")
        self.resizable(0, 0)
        app.destroy()
        self.Financial_Interface()

    # 转入收入页
    def shouruxian(self):
        self.frame_chu.pack_forget()
        self.frame_ru.pack(side=tk.TOP)
        self.frame_ru.config(bg="#e1f5c4")
        self.button_ru.pack_forget()

    # 转入支出页
    def zhichuxian(self):
        self.frame_ru.pack_forget()
        self.frame_chu.pack(side=tk.TOP)
        self.frame_chu.config(bg="#e1f5c4")
        self.button_chu.pack_forget()

    # 收入录入按钮函数
    def luru_shouru(self):
        data_shouru = self.txt_data.get()
        print(data_shouru)
        money_shouru = self.txt_money.get()
        print(money_shouru)
        minxi_shouru = self.txt_minxi.get()
        print(minxi_shouru)
        remarks_shouru = self.txt_remarks.get('0.0', 'end')
        print(remarks_shouru)
        if data_shouru != '' and money_shouru != '' and minxi_shouru != '':
            wb = openpyxl.load_workbook('Z.xlsx')
            ws = wb.worksheets[1]
            ws.append((data_shouru, money_shouru,
                       minxi_shouru, remarks_shouru))
            wb.save('Z.xlsx')
            self.txt_money.delete(0, tk.END)
            self.txt_minxi.delete(0, tk.END)
            self.txt_remarks.delete('1.0', 'end')
        else:
            self.frame_ru1.pack()
            self.lab_shuruxiang.config(
                text='*为必填选项', font=('微软雅黑', '18'), fg='red')
            self.frame_ru1.after(3000, self.frame_ru1.pack_forget)
            # self.frame_ru1.pack_forget()

    # 支出录入按钮函数

    def luru_zhichu(self):
        data_shouru = self.txt_datachu.get()
        print(data_shouru)
        money_shouru = self.txt_moneychu.get()
        print(money_shouru)
        minxi_shouru = self.txt_minxichu.get()
        print(minxi_shouru)
        remarks_shouru = self.txt_remarkschu.get('0.0', 'end')
        print(remarks_shouru)
        if data_shouru != '' and money_shouru != '' and minxi_shouru != '':
            wb = openpyxl.load_workbook('Z.xlsx')
            ws = wb.worksheets[2]
            ws.append((data_shouru, money_shouru,
                       minxi_shouru, remarks_shouru))
            wb.save('Z.xlsx')

            self.txt_datachu.delete(0, tk.END)
            self.txt_moneychu.delete(0, tk.END)
            self.txt_minxichu.delete(0, tk.END)
            self.txt_remarkschu.delete('1.0', 'end')
        else:
            self.frame_ru1.pack()
            self.lab_shuruxiang.config(
                text='*为必填选项', font=('微软雅黑', '18'), fg='red')
            self.frame_ru1.after(3000, self.frame_ru1.pack_forget)

    # 收入页显示

    def shouru(self):
        self.frame_chu.pack_forget()
        self.frame_ru.pack(side=tk.TOP)
        self.frame_ru.config(bg="#e1f5c4")
        self.button_ru.pack_forget()

        self.lab_ru = tk.Label(self.frame_ru, text="收入明细录入", font=(
            "微软雅黑", 24), height=2, relief='groove', bg='#f1f5c4')
        self.lab_ru.pack(fill=tk.X)

        self.btn_goright = tk.Button(self.lab_ru, text="》》", width="4", height="2", relief='groove', bg='#f1f5c4',
                                     command=self.zhichuxian)
        self.btn_goright.pack(side=tk.RIGHT)

        self.lbl_data = tk.Label(self.frame_ru, text="*日期", anchor=tk.W)
        self.lbl_data.pack(fill=tk.X, padx=10)

        alls = datetime.date.today()
        self.txt_data = tk.Entry(self.frame_ru, textvariable=alls)
        self.txt_data.insert(0, alls)
        self.txt_data.pack(fill=tk.X, padx=10)

        self.lbl_money = tk.Label(self.frame_ru, text="*金额", anchor=tk.W)
        self.lbl_money.pack(fill=tk.X, padx=10)

        self.txt_money = tk.Entry(self.frame_ru)
        self.txt_money.pack(fill=tk.X, padx=10)

        self.lbl_minxi = tk.Label(self.frame_ru, text="*明细", anchor=tk.W)
        self.lbl_minxi.pack(fill=tk.X, padx=10)

        self.txt_minxi = tk.Entry(self.frame_ru)
        self.txt_minxi.pack(fill=tk.X, padx=10)

        self.lbl_remarks = tk.Label(self.frame_ru, text="备注", anchor=tk.W)
        self.lbl_remarks.pack(fill=tk.X, padx=10)

        self.txt_remarks = tk.Text(self.frame_ru, width=40, height=10)
        self.txt_remarks.pack(padx=10)

        self.btn_Input = tk.Button(
            self.frame_ru, text="录入", width="16", command=self.luru_shouru)
        self.btn_Input.pack(side=tk.LEFT, pady=10, padx=10)

        self.btn_check = tk.Button(
            self.frame_ru, text="查询", width="16", command=chaxunhanshu.chaxunhanshu)
        self.btn_check.pack(side=tk.RIGHT, pady=10, padx=10)

    #! 支出页显示
    def zhichu(self):
        self.frame_ru.pack_forget()
        self.frame_chu.pack(side=tk.TOP)
        self.frame_chu.config(bg="#e1f5c4")
        self.button_chu.pack_forget()

        self.lab_chu = tk.Label(self.frame_chu, text="支出明细录入", font=("微软雅黑", 24), height=2, relief='groove',
                                bg='#f1f5c4')
        self.lab_chu.pack(fill=tk.X)

        self.btn_goleft = tk.Button(self.lab_chu, text="《《", width="4", height="2", relief='groove', bg='#f1f5c4',
                                    command=self.shouruxian)
        self.btn_goleft.pack(side=tk.LEFT)

        self.lbl_datachu = tk.Label(self.frame_chu, text="日期", anchor=tk.W)
        self.lbl_datachu.pack(fill=tk.X, padx=10)

        self.txt_datachu = tk.Entry(self.frame_chu)
        self.txt_datachu.pack(fill=tk.X, padx=10)

        self.lbl_moneychu = tk.Label(self.frame_chu, text="金额", anchor=tk.W)
        self.lbl_moneychu.pack(fill=tk.X, padx=10)

        self.txt_moneychu = tk.Entry(self.frame_chu)
        self.txt_moneychu.pack(fill=tk.X, padx=10)

        self.lbl_minxichu = tk.Label(self.frame_chu, text="明细", anchor=tk.W)
        self.lbl_minxichu.pack(fill=tk.X, padx=10)

        self.txt_minxichu = tk.Entry(self.frame_chu)
        self.txt_minxichu.pack(fill=tk.X, padx=10)

        self.lbl_remarkschu = tk.Label(self.frame_chu, text="备注", anchor=tk.W)
        self.lbl_remarkschu.pack(fill=tk.X, padx=10)

        self.txt_remarkschu = tk.Text(self.frame_chu, width=40, height=10)
        self.txt_remarkschu.pack(padx=10)

        self.btn_Inputchu = tk.Button(
            self.frame_chu, text="录入", width="16", command=self.luru_zhichu)
        self.btn_Inputchu.pack(side=tk.LEFT, pady=10, padx=10)

        self.btn_checkchu = tk.Button(self.frame_chu, text="查询", width="16")
        self.btn_checkchu.pack(side=tk.RIGHT, pady=10, padx=10)

    # 从登录页转到录入页
    def Financial_Interface(self):
        self.frame_ru = tk.Frame(self, bg="#e1f5c4")
        self.frame_ru.pack(side=tk.LEFT)
        self.frame_ru1 = tk.Frame(self, bg="#e1fec4")
        self.frame_ru1.pack_forget()
        self.lab_shuruxiang = tk.Label(self.frame_ru1)
        self.lab_shuruxiang.pack()

        self.button_ru = tk.Button(self.frame_ru, text='收入', bg="#ede574", wraplength=1, width=4, height=4,
                                   font=("微软雅黑", 34), relief='groove',
                                   activebackground='red', command=self.shouru)
        self.button_ru.pack(padx=10, pady=13, anchor=tk.CENTER)

        self.frame_chu = tk.Frame(self)
        self.frame_chu.pack(side=tk.RIGHT)
        self.button_chu = tk.Button(self.frame_chu, text='支出', bg="#ede574", wraplength=1, width=4, height=4,
                                    font=("微软雅黑", 34), relief='groove',
                                    activebackground='red', command=self.zhichu)
        self.button_chu.pack(padx=10, pady=13, anchor=tk.CENTER)

        self.shouru()
        self.zhichu()


# 登录界面
class Application(tk.Tk):

    # 初始化
    def __init__(self):
        super().__init__()
        self.title("谢立虎个人财务系统登陆")
        self.geometry("+800+250")
        self.resizable(0, 0)
        self.init_widgets()

    # 登录界面中点击登录函数
    def denglu(self):
        # self.frame1.pack(padx=20, pady=5, fill=tk.X)
        try:
            username = self.txt_username.get()
            password = self.txt_password.get()

            excel = load_workbook('Z.xlsx')
            sheet = excel.get_sheet_by_name('C')

            excel_username = sheet.cell(row=1, column=1).value
            excel_password = sheet.cell(row=1, column=2).value
            if username == excel_username and password == excel_password:
                Financial_system()
            else:
                self.frame1.pack()
                self.lbl_baocuo.config(
                    text='您输入的账户或密码不正确', font=('微软雅黑', '18'), fg='red')
                self.txt_username.delete(0, tk.END)
                self.txt_password.delete(0, tk.END)
                self.frame1.after(2000, self.frame1.pack_forget)
        except Exception as e:
            self.frame1.pack()
            self.lbl_baocuo.config(text='请先注册！', font=('微软雅黑', '18'), fg='red')
            self.txt_username.delete(0, tk.END)
            self.txt_password.delete(0, tk.END)
            self.frame1.after(2000, self.frame1.pack_forget)

    # 生成登录界面

    def init_widgets(self):
        photo = ImageTk.PhotoImage(Image.open("D:\python\PYTHON\JIZHANG\cw.jpg"))
        self.lbl_img = tk.Label(self, image=photo, width=300, height=50)
        self.lbl_img.image = photo
        self.lbl_img.pack()

        self.frame = tk.Frame(self)

        self.lbl_username = tk.Label(self.frame, text="用户名", anchor=tk.W)
        self.lbl_username.pack(fill=tk.X)

        self.txt_username = tk.Entry(self.frame)
        self.txt_username.pack(fill=tk.X)

        self.lbl_password = tk.Label(self.frame, text="密码", anchor=tk.W)
        self.lbl_password.pack(fill=tk.X)

        self.txt_password = tk.Entry(self.frame, show="*")
        self.txt_password.pack(fill=tk.X)

        self.btn_login = tk.Button(
            self.frame, text="登录", width="16", command=self.denglu)
        self.btn_login.pack(side=tk.LEFT, pady=10)

        self.btn_cancel = tk.Button(
            self.frame, text="注册", width="16", command=zhuce)
        self.btn_cancel.pack(side=tk.RIGHT, pady=10)

        self.frame.pack(padx=20, pady=5, fill=tk.X)

        self.frame1 = tk.Frame(self)

        self.lbl_baocuo = tk.Label(self.frame1, text="", anchor=tk.CENTER)
        self.lbl_baocuo.pack()

        self.frame1.pack_forget()


# 收入查询界面


# 主函数
if __name__ == '__main__':
    app = Application()

    app.mainloop()
